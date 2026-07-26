from datetime import date
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook

from iconomics.config import find_config_dir
from iconomics.money import Money
from iconomics.workbook import MissingColumn, Sheet, load, write


def data_raw():
    return find_config_dir().parent / "data" / "raw"


def make_file(tmp_path, headers, rows, name="in.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    path = tmp_path / name
    workbook.save(path)
    return path


def test_cyrillic_headers_map_to_canonical_fields(tmp_path):
    path = make_file(
        tmp_path,
        ["Дата", "Контрагент", "Сума без ДДС", "ДДС"],
        [["01.02.2026", "Алфа ООД", "100,00", "20,00"]],
    )
    ledger = load(path)
    assert len(ledger.rows) == 1
    row = ledger.rows[0]
    assert row.date == date(2026, 2, 1)
    assert row.counterparty == "Алфа ООД"
    assert row.amount_net == Money(Decimal("100.00"), "EUR")
    assert row.vat_amount == Money(Decimal("20.00"), "EUR")


def test_alternate_header_spelling_also_maps(tmp_path):
    path = make_file(
        tmp_path,
        ["Дата", "Партньор", "Сума без ДДС"],
        [["01.02.2026", "Алфа ООД", "100,00"]],
    )
    assert load(path).rows[0].counterparty == "Алфа ООД"


def test_source_row_is_the_original_spreadsheet_row(tmp_path):
    path = make_file(
        tmp_path,
        ["Дата", "Контрагент", "Сума без ДДС"],
        [["01.02.2026", "Алфа ООД", "1,00"], ["02.02.2026", "Бета ЕООД", "2,00"]],
    )
    # Header is row 1, so data starts at row 2.
    assert [row.source_row for row in load(path).rows] == [2, 3]


def test_money_is_always_decimal_backed(tmp_path):
    path = make_file(
        tmp_path,
        ["Дата", "Контрагент", "Сума без ДДС"],
        [["01.02.2026", "Алфа ООД", 100.5]],
    )
    amount = load(path).rows[0].amount_net
    assert isinstance(amount.amount, Decimal)
    assert not isinstance(amount.amount, float)


def test_currency_defaults_by_date_when_no_currency_column(tmp_path):
    path = make_file(
        tmp_path,
        ["Дата", "Контрагент", "Сума без ДДС"],
        [["31.12.2025", "Алфа ООД", "100,00"], ["01.01.2026", "Алфа ООД", "100,00"]],
    )
    rows = load(path).rows
    assert rows[0].currency == "BGN"
    assert rows[1].currency == "EUR"


def test_explicit_currency_column_overrides_the_date_default(tmp_path):
    path = make_file(
        tmp_path,
        ["Дата", "Контрагент", "Сума без ДДС", "Валута"],
        [["01.02.2026", "Алфа ООД", "100,00", "BGN"]],
    )
    assert load(path).rows[0].currency == "BGN"


def test_bad_amount_becomes_a_problem_and_does_not_abort(tmp_path):
    path = make_file(
        tmp_path,
        ["Дата", "Контрагент", "Сума без ДДС"],
        [["01.02.2026", "Алфа ООД", "n/a"], ["02.02.2026", "Бета ЕООД", "5,00"]],
    )
    ledger = load(path)
    assert len(ledger.rows) == 1
    assert len(ledger.problems) == 1
    problem = ledger.problems[0]
    assert problem.source_row == 2
    assert problem.field == "amount_net"
    assert "n/a" in problem.raw


def test_blank_date_becomes_a_problem_described_in_plain_words(tmp_path):
    path = make_file(
        tmp_path, ["Дата", "Контрагент", "Сума без ДДС"], [[None, "Гама АД", "9,00"]]
    )
    ledger = load(path)
    assert ledger.rows == []
    problem = ledger.problems[0]
    assert problem.field == "date"
    assert problem.raw == "(blank)"
    assert problem.reason == "empty cell"


def test_missing_required_column_aborts_before_output(tmp_path):
    path = make_file(tmp_path, ["Дата", "Контрагент"], [["01.02.2026", "Алфа ООД"]])
    with pytest.raises(MissingColumn, match="amount_net"):
        load(path)


def test_unmapped_columns_are_preserved_not_dropped(tmp_path):
    path = make_file(
        tmp_path,
        ["Дата", "Контрагент", "Сума без ДДС", "Вътрешен код"],
        [["01.02.2026", "Алфа ООД", "1,00", "XYZ-1"]],
    )
    ledger = load(path)
    assert ledger.unmapped_headers == ["Вътрешен код"]
    assert ledger.rows[0].extra["Вътрешен код"] == "XYZ-1"


def test_fully_blank_rows_are_skipped_silently(tmp_path):
    path = make_file(
        tmp_path,
        ["Дата", "Контрагент", "Сума без ДДС"],
        [["01.02.2026", "Алфа ООД", "1,00"], [None, None, None]],
    )
    ledger = load(path)
    assert len(ledger.rows) == 1
    assert ledger.problems == []


def test_all_sample_files_load():
    for name in ("ledger-2025-12.xlsx", "ledger-2026-01.xlsx", "ledger-2026-02.xlsx"):
        ledger = load(data_raw() / name)
        assert ledger.rows, f"{name} produced no rows"


def test_write_produces_readable_multi_sheet_output(tmp_path):
    path = tmp_path / "out.xlsx"
    write(
        path,
        {
            "Clean": Sheet(columns=["A", "B"], rows=[[1, "x"], [2, "y"]]),
            "Exceptions": Sheet(columns=["Row", "Reason"], rows=[[7, "bad amount"]]),
        },
    )
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Clean", "Exceptions"]
    assert [cell.value for cell in workbook["Clean"][1]] == ["A", "B"]
    assert workbook["Exceptions"]["A2"].value == 7


def test_write_serializes_money_and_decimal_as_numbers(tmp_path):
    path = tmp_path / "out.xlsx"
    write(
        path,
        {"S": Sheet(columns=["Amount"], rows=[[Money(Decimal("12.34"), "EUR")]])},
    )
    assert load_workbook(path)["S"]["A2"].value == 12.34


def test_write_applies_money_number_formats(tmp_path):
    path = tmp_path / "out.xlsx"
    write(
        path,
        {
            "S": Sheet(
                columns=["Net", "Note"],
                rows=[[Money(Decimal("410.00"), "EUR"), "x"]],
                number_formats={"Net": "0.00"},
            )
        },
    )
    sheet = load_workbook(path)["S"]
    assert sheet["A2"].number_format == "0.00"
    assert sheet["B2"].number_format == "General"


def test_write_freezes_the_header_row(tmp_path):
    path = tmp_path / "out.xlsx"
    write(path, {"S": Sheet(columns=["A"], rows=[[1]])})
    assert load_workbook(path)["S"].freeze_panes == "A2"
