"""The only module that touches .xlsx files.

Two boundary guarantees:
  * every monetary value leaving load() is Decimal-backed, never float
  * every row and problem carries source_row, the 1-indexed original row
"""

from dataclasses import dataclass, field
from datetime import date as Date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from iconomics.config import load_header_aliases
from iconomics.money import Money, default_currency_for
from iconomics.parsing import (
    UnparseableAmount,
    UnparseableDate,
    normalize_direction,
    normalize_header,
    parse_amount,
    parse_date,
)

REQUIRED_FIELDS = ("date", "counterparty", "amount_net")
#: A bank statement has no counterparty column — only a description.
STATEMENT_FIELDS = ("date", "amount_net")
#: A trial balance has no dates at all.
TRIAL_BALANCE_FIELDS = ("account",)
_HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")


class MissingColumn(RuntimeError):
    """Raised when the file lacks a column the toolkit cannot work without."""


@dataclass(frozen=True)
class Row:
    source_row: int
    date: Date
    counterparty: str
    description: str
    amount_net: Money
    currency: str
    vat_number: str | None = None
    vat_amount: Money | None = None
    vat_rate: Decimal | None = None
    account: str | None = None
    direction: str | None = None
    extra: dict[str, object] = field(default_factory=dict)


@dataclass(frozen=True)
class Problem:
    source_row: int
    field: str
    raw: str
    reason: str


@dataclass(frozen=True)
class Ledger:
    rows: list[Row]
    problems: list[Problem]
    unmapped_headers: list[str]
    source_path: Path


@dataclass
class Sheet:
    columns: list[str]
    rows: list[list[object]]
    #: Column name -> Excel number format, e.g. {"Net": "0.00"}. Applied to the
    #: whole column below the header so money reads as money.
    number_formats: dict[str, str] | None = None


def _is_blank(value: object) -> bool:
    return value is None or str(value).strip() == ""


def _describe_raw(value: object) -> str:
    """Render a rejected cell value for the Exceptions sheet.

    A blank cell says so in words: "None" is a programmer's answer to an
    accountant's question.
    """
    return "(blank)" if _is_blank(value) else str(value)


def _map_columns(header_cells, aliases):
    """Return (field -> column index) and the list of unmapped header labels."""
    mapping: dict[str, int] = {}
    unmapped: list[str] = []
    for index, raw in enumerate(header_cells):
        if _is_blank(raw):
            continue
        canonical = aliases.get(normalize_header(raw))
        if canonical is None:
            unmapped.append(str(raw))
        elif canonical not in mapping:
            mapping[canonical] = index
    return mapping, unmapped


def _cell(values, mapping, name):
    index = mapping.get(name)
    if index is None or index >= len(values):
        return None
    return values[index]


def load(
    path: Path,
    aliases: dict[str, str] | None = None,
    required: tuple[str, ...] = REQUIRED_FIELDS,
) -> Ledger:
    """Read a spreadsheet into canonical rows plus a list of row-level problems.

    ``required`` names the columns the file must have. Ledgers need a
    counterparty; bank statements do not, so callers pass a narrower set
    rather than each growing its own loader.
    """
    path = Path(path)
    resolved_aliases = aliases if aliases is not None else load_header_aliases()

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        raise MissingColumn(f"{path} is empty")

    headers = all_rows[0]
    mapping, unmapped = _map_columns(headers, resolved_aliases)
    for required_field in required:
        if required_field not in mapping:
            raise MissingColumn(
                f"{path} has no column mapping to {required_field!r}; "
                "add an alias in config/headers.yaml"
            )

    rows: list[Row] = []
    problems: list[Problem] = []

    for offset, values in enumerate(all_rows[1:], start=2):
        if all(_is_blank(value) for value in values):
            continue

        raw_date = _cell(values, mapping, "date")
        try:
            row_date = parse_date(raw_date)
        except UnparseableDate as exc:
            problems.append(Problem(offset, "date", _describe_raw(raw_date), str(exc)))
            continue

        raw_net = _cell(values, mapping, "amount_net")
        try:
            net = parse_amount(raw_net)
        except UnparseableAmount as exc:
            problems.append(Problem(offset, "amount_net", _describe_raw(raw_net), str(exc)))
            continue

        currency_raw = _cell(values, mapping, "currency")
        currency = (
            str(currency_raw).strip().upper()
            if not _is_blank(currency_raw)
            else default_currency_for(row_date)
        )

        vat_raw = _cell(values, mapping, "vat_amount")
        vat_amount = None
        if not _is_blank(vat_raw):
            try:
                vat_amount = Money(parse_amount(vat_raw), currency)
            except UnparseableAmount as exc:
                problems.append(Problem(offset, "vat_amount", _describe_raw(vat_raw), str(exc)))
                continue

        rate_raw = _cell(values, mapping, "vat_rate")
        vat_rate = None
        if not _is_blank(rate_raw):
            try:
                vat_rate = parse_amount(rate_raw)
            except UnparseableAmount as exc:
                problems.append(Problem(offset, "vat_rate", _describe_raw(rate_raw), str(exc)))
                continue

        extra = {
            str(headers[index]): values[index]
            for index in range(len(headers))
            if not _is_blank(headers[index])
            and str(headers[index]) in unmapped
            and index < len(values)
        }

        counterparty_raw = _cell(values, mapping, "counterparty")
        vat_number_raw = _cell(values, mapping, "vat_number")
        account_raw = _cell(values, mapping, "account")
        description_raw = _cell(values, mapping, "description")
        direction = normalize_direction(_cell(values, mapping, "direction"))

        rows.append(
            Row(
                source_row=offset,
                date=row_date,
                counterparty=str(counterparty_raw or "").strip(),
                description=str(description_raw or "").strip(),
                amount_net=Money(net, currency),
                currency=currency,
                vat_number=str(vat_number_raw).strip() if vat_number_raw else None,
                vat_amount=vat_amount,
                vat_rate=vat_rate,
                account=str(account_raw).strip() if account_raw else None,
                direction=direction,
                extra=extra,
            )
        )

    return Ledger(
        rows=rows, problems=problems, unmapped_headers=unmapped, source_path=path
    )


@dataclass(frozen=True)
class TrialLine:
    source_row: int
    account: str
    name: str
    debit: Money
    credit: Money


@dataclass(frozen=True)
class TrialBalance:
    lines: list[TrialLine]
    problems: list[Problem]
    source_path: Path


def load_trial_balance(
    path: Path, aliases: dict[str, str] | None = None, currency: str = "EUR"
) -> TrialBalance:
    """Read a trial balance: account code, name, debit and credit balances.

    A separate loader because a trial balance has no dates and no counterparties
    — forcing it through load() would mean weakening the ledger contract for
    every caller.
    """
    path = Path(path)
    resolved = aliases if aliases is not None else load_header_aliases()

    sheet = load_workbook(path, data_only=True).active
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        raise MissingColumn(f"{path} is empty")

    headers = all_rows[0]
    mapping, unmapped = _map_columns(headers, resolved)
    if "account" not in mapping:
        raise MissingColumn(
            f"{path} has no column mapping to 'account'; add an alias in "
            "config/headers.yaml"
        )
    if "debit" not in mapping or "credit" not in mapping:
        raise MissingColumn(
            f"{path} needs both a debit and a credit column; add aliases in "
            "config/headers.yaml"
        )

    lines: list[TrialLine] = []
    problems: list[Problem] = []
    zero = Money(Decimal("0.00"), currency)

    for offset, values in enumerate(all_rows[1:], start=2):
        if all(_is_blank(value) for value in values):
            continue

        account_raw = _cell(values, mapping, "account")
        if _is_blank(account_raw):
            problems.append(Problem(offset, "account", "(blank)", "empty cell"))
            continue

        amounts = {}
        failed = False
        for side in ("debit", "credit"):
            raw = _cell(values, mapping, side)
            if _is_blank(raw):
                amounts[side] = zero
                continue
            try:
                amounts[side] = Money(parse_amount(raw), currency)
            except UnparseableAmount as exc:
                problems.append(Problem(offset, side, _describe_raw(raw), str(exc)))
                failed = True
                break
        if failed:
            continue

        name_raw = _cell(values, mapping, "description")
        lines.append(
            TrialLine(
                source_row=offset,
                account=str(account_raw).strip(),
                name=str(name_raw or "").strip(),
                debit=amounts["debit"],
                credit=amounts["credit"],
            )
        )

    return TrialBalance(lines=lines, problems=problems, source_path=path)


def _serialize(value: object) -> object:
    """Flatten a value into something openpyxl can store.

    This is the one place a float is permitted: the value is leaving the system
    into a spreadsheet cell, after all arithmetic is complete, and is never
    read back for computation.
    """
    if isinstance(value, Money):
        return float(value.amount)
    if isinstance(value, Decimal):
        return float(value)
    # An empty string and a blank cell must not be different things: openpyxl
    # reads "" back as None while exceljs reads it back as "", which would make
    # the two implementations disagree on cells that are both simply empty.
    if value == "":
        return None
    return value


def write(path: Path, sheets: dict[str, Sheet]) -> None:
    """Write formatted, frozen-header sheets in the given order."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    for name, spec in sheets.items():
        sheet = workbook.create_sheet(title=name)
        sheet.append(spec.columns)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(vertical="center")
        for row in spec.rows:
            sheet.append([_serialize(value) for value in row])

        formats = spec.number_formats or {}
        for column_name, number_format in formats.items():
            if column_name not in spec.columns:
                continue
            letter = get_column_letter(spec.columns.index(column_name) + 1)
            for row_index in range(2, len(spec.rows) + 2):
                sheet[f"{letter}{row_index}"].number_format = number_format

        sheet.freeze_panes = "A2"
        for index, column in enumerate(spec.columns, start=1):
            widths = [len(str(column))] + [
                len(str(_serialize(row[index - 1])))
                for row in spec.rows
                if index - 1 < len(row)
            ]
            sheet.column_dimensions[get_column_letter(index)].width = min(
                max(widths) + 2, 48
            )

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
