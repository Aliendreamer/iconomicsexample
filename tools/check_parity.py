#!/usr/bin/env python3
"""Check that the Python and JavaScript implementations agree exactly.

Runs both CLIs over every sample ledger and compares the resulting workbooks
cell by cell, plus the stdout summaries. This is what makes two implementations
safe to keep: if one drifts, this says so instead of shipping a wrong figure.

Not a test framework — just a script.

    python tools/check_parity.py

Exit code 0 if the two agree, 1 if they do not.
"""

import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DATA_RAW = ROOT / "data" / "raw"
JS_CLI = ROOT / "js" / "bin" / "iconomics.js"
PY = ROOT / ".venv" / "bin" / "python"


def run(command, out_dir):
    result = subprocess.run(
        command + ["--out", str(out_dir)],
        capture_output=True,
        text=True,
        cwd=ROOT,
    )
    if result.returncode != 0:
        raise SystemExit(f"command failed: {' '.join(command)}\n{result.stderr}")
    return result.stdout


def read_cells(path):
    workbook = load_workbook(path)
    return {
        name: [list(row) for row in workbook[name].iter_rows(values_only=True)]
        for name in workbook.sheetnames
    }


def compare(sample, problems):
    stem = sample.stem
    with tempfile.TemporaryDirectory() as tmp:
        py_dir = Path(tmp) / "py"
        js_dir = Path(tmp) / "js"

        py_out = run(
            [str(PY), "-m", "iconomics", "cleanup", "--in", str(sample)], py_dir
        )
        js_out = run(
            ["node", str(JS_CLI), "cleanup", "--in", str(sample)], js_dir
        )

        # Every summary line except the output path, which names a temp dir.
        py_lines = py_out.strip().splitlines()[:-1]
        js_lines = js_out.strip().splitlines()[:-1]
        if py_lines != js_lines:
            problems.append(f"{stem}: stdout summaries differ")
            problems.append(f"    python: {py_lines}")
            problems.append(f"    node:   {js_lines}")

        py_cells = read_cells(py_dir / f"{stem}-clean.xlsx")
        js_cells = read_cells(js_dir / f"{stem}-clean.xlsx")

        if list(py_cells) != list(js_cells):
            problems.append(
                f"{stem}: sheet names differ — {list(py_cells)} vs {list(js_cells)}"
            )
            return

        for sheet_name in py_cells:
            py_rows = py_cells[sheet_name]
            js_rows = js_cells[sheet_name]
            if len(py_rows) != len(js_rows):
                problems.append(
                    f"{stem}[{sheet_name}]: {len(py_rows)} rows vs {len(js_rows)}"
                )
            for index, (py_row, js_row) in enumerate(zip(py_rows, js_rows), start=1):
                if py_row != js_row:
                    problems.append(f"{stem}[{sheet_name}] row {index}:")
                    problems.append(f"    python: {py_row}")
                    problems.append(f"    node:   {js_row}")


#: The three later workflows, as (label, argument builder, output filename).
#: Each entry runs both CLIs and diffs every sheet of the resulting workbook.
OTHER_COMMANDS = (
    (
        "vat-return",
        lambda: [
            "vat-return",
            "--in",
            str(DATA_RAW / "journal-2026-03.xlsx"),
        ],
        "journal-2026-03-vat-return.xlsx",
    ),
    (
        "statements",
        lambda: [
            "statements",
            "--in",
            str(DATA_RAW / "trial-balance-2026-03.xlsx"),
            "--prior",
            str(DATA_RAW / "trial-balance-2025-12.xlsx"),
            "--prior-currency",
            "BGN",
        ],
        "trial-balance-2026-03-statements.xlsx",
    ),
    (
        "reconcile",
        lambda: [
            "reconcile",
            "--bank",
            str(DATA_RAW / "bank-2026-03.xlsx"),
            "--ledger",
            str(DATA_RAW / "ledger-2026-03.xlsx"),
        ],
        "bank-2026-03-reconciliation.xlsx",
    ),
)


def compare_other_commands(problems):
    """Diff vat-return, statements and reconcile across both implementations."""
    for label, arguments, filename in OTHER_COMMANDS:
        with tempfile.TemporaryDirectory() as tmp:
            py_dir = Path(tmp) / "py"
            js_dir = Path(tmp) / "js"

            py_out = run([str(PY), "-m", "iconomics"] + arguments(), py_dir)
            js_out = run(["node", str(JS_CLI)] + arguments(), js_dir)

            py_lines = py_out.strip().splitlines()[:-1]
            js_lines = js_out.strip().splitlines()[:-1]
            if py_lines != js_lines:
                problems.append(f"{label}: stdout summaries differ")
                problems.append(f"    python: {py_lines}")
                problems.append(f"    node:   {js_lines}")

            py_cells = read_cells(py_dir / filename)
            js_cells = read_cells(js_dir / filename)

            if list(py_cells) != list(js_cells):
                problems.append(
                    f"{label}: sheet names differ — {list(py_cells)} vs {list(js_cells)}"
                )
                continue

            for sheet_name in py_cells:
                py_rows = py_cells[sheet_name]
                js_rows = js_cells[sheet_name]
                if len(py_rows) != len(js_rows):
                    problems.append(
                        f"{label}[{sheet_name}]: {len(py_rows)} rows vs {len(js_rows)}"
                    )
                for index, (py_row, js_row) in enumerate(
                    zip(py_rows, js_rows), start=1
                ):
                    if py_row != js_row:
                        problems.append(f"{label}[{sheet_name}] row {index}:")
                        problems.append(f"    python: {py_row}")
                        problems.append(f"    node:   {js_row}")

        print(f"checked {label}")


#: (kind, complexity) pairs to check. The ledger is checked at all three
#: complexity levels; the other kinds at `nasty`, which exercises the most code.
GENERATE_CASES = (
    ("ledger", "clean"),
    ("ledger", "messy"),
    ("ledger", "nasty"),
    ("journal", "nasty"),
    ("trial-balance", "nasty"),
    ("bank", "nasty"),
)


def compare_generated(problems):
    """The generator must also agree — it is deterministic in both languages."""
    for kind, complexity in GENERATE_CASES:
        with tempfile.TemporaryDirectory() as tmp:
            py_file = Path(tmp) / "py.xlsx"
            js_file = Path(tmp) / "js.xlsx"

            for command, target in (
                ([str(PY), "-m", "iconomics"], py_file),
                (["node", str(JS_CLI)], js_file),
            ):
                result = subprocess.run(
                    command
                    + [
                        "generate",
                        "--kind",
                        kind,
                        "--rows",
                        "40",
                        "--complexity",
                        complexity,
                        "--out",
                        str(target),
                    ],
                    capture_output=True,
                    text=True,
                    cwd=ROOT,
                )
                if result.returncode != 0:
                    raise SystemExit(f"generate failed ({kind}/{complexity}):\n{result.stderr}")

            py_cells = read_cells(py_file)["Ledger"]
            js_cells = read_cells(js_file)["Ledger"]

            if len(py_cells) != len(js_cells):
                problems.append(
                    f"generate/{kind}/{complexity}: {len(py_cells)} rows vs {len(js_cells)}"
                )
            for index, (py_row, js_row) in enumerate(zip(py_cells, js_cells), start=1):
                if py_row != js_row:
                    problems.append(f"generate/{kind}/{complexity} row {index}:")
                    problems.append(f"    python: {py_row}")
                    problems.append(f"    node:   {js_row}")

        print(f"checked generate --kind {kind} --complexity {complexity}")


def main():
    if not PY.is_file():
        raise SystemExit("no .venv — run: uv venv --python 3.12 && uv pip install -e py/")
    if shutil.which("node") is None:
        raise SystemExit("node not found on PATH")
    if not (ROOT / "js" / "node_modules").is_dir():
        raise SystemExit("js dependencies missing — run: npm --prefix js install")

    # Only the ledger-shaped files: cleanup requires a counterparty column, which
    # a bank statement and a trial balance deliberately do not have. Those two are
    # exercised by compare_other_commands instead.
    samples = sorted(DATA_RAW.glob("ledger-*.xlsx")) + sorted(
        DATA_RAW.glob("journal-*.xlsx")
    )
    if not samples:
        raise SystemExit("no sample data — run: python tools/make_sample_data.py")

    problems = []
    for sample in samples:
        compare(sample, problems)
        print(f"checked {sample.name}")

    compare_other_commands(problems)
    compare_generated(problems)

    if problems:
        print("\nPARITY FAILURES:")
        for line in problems:
            print(f"  {line}")
        return 1

    print(
        f"\nparity OK — both implementations agree on {len(samples)} cleanup runs, "
        f"{len(OTHER_COMMANDS)} workflow commands, and {len(GENERATE_CASES)} generated files"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
